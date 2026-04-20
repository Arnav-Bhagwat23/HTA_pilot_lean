from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from hta_pipeline.excel_export import write_extraction_excel


def _sample_record():
    return {
        "schema_version": "1.0",
        "document_set": {
            "product_name": "Jemperli",
            "indication": None,
            "country": "United Kingdom",
            "documents_considered": [
                {
                    "document_id": "smc_uk::sample",
                    "title": "dostarlimab (Jemperli)",
                    "source_id": "smc_uk",
                    "source_name": "Scottish Medicines Consortium",
                    "document_type": "HTA guidance",
                    "format": "pdf",
                    "document_url": "https://example.test/sample.pdf",
                    "local_file_path": "downloads/sample.pdf",
                    "event_date": "2026-01-01",
                    "publication_date": "2026-01-01",
                    "revision_date": None,
                    "is_latest_version": True,
                    "timeline_priority": 1,
                    "match_term": "Jemperli",
                    "match_confidence": "title_match",
                    "processing_order": 1,
                }
            ],
        },
        "progressive_fill": {
            "strategy": "newest_explicit_then_backfill_then_controlled_inference",
            "allow_final_inference": True,
            "overwrite_populated_fields": False,
            "missing_field_policy": "leave_null_when_not_supported",
        },
        "hta_results": {
            "hta_outcome": {
                "value": "Not recommended for use within NHSScotland.",
                "fill_method": "explicit_latest",
                "source_document_id": "smc_uk::sample",
                "source_document_title": "dostarlimab (Jemperli)",
                "source_document_url": "https://example.test/sample.pdf",
                "source_document_date": "2026-01-01",
                "source_page": "1",
                "evidence_snippet": "not recommended for use within NHSScotland",
                "confidence": "high",
                "warnings": [],
            },
            "reimbursed_population": {
                "value": None,
                "fill_method": "not_found",
                "source_document_id": None,
                "source_document_title": None,
                "source_document_url": None,
                "source_document_date": None,
                "source_page": None,
                "evidence_snippet": None,
                "confidence": "unknown",
                "warnings": ["Field was not found after progressive extraction."],
            },
        },
        "traceability": {
            "created_at": "2026-04-20T10:00:00+00:00",
            "updated_at": "2026-04-20T10:01:00+00:00",
            "extraction_model": "test-model",
            "extraction_status": "partial",
            "warnings": [],
            "audit_log": [
                {
                    "timestamp": "2026-04-20T10:00:30+00:00",
                    "action": "explicit_latest",
                    "document_id": "smc_uk::sample",
                    "fields_attempted": ["hta_outcome", "reimbursed_population"],
                    "fields_filled": ["hta_outcome"],
                    "notes": None,
                }
            ],
        },
    }


class ExcelExportTests(unittest.TestCase):
    def test_write_extraction_excel_creates_six_review_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            destination = Path(temp_dir) / "export.xlsx"
            write_extraction_excel(_sample_record(), destination)
            workbook = load_workbook(destination)

        self.assertEqual(
            workbook.sheetnames,
            [
                "HTA Results",
                "Documents Considered",
                "Extraction Audit Log",
                "Missing Fields & Warnings",
                "Run Metadata",
                "Source URLs",
            ],
        )
        self.assertEqual(workbook["HTA Results"]["A1"].value, "product_name")
        self.assertEqual(
            workbook["HTA Results"]["F2"].value,
            "Not recommended for use within NHSScotland.",
        )
        self.assertEqual(
            workbook["Missing Fields & Warnings"]["F2"].value,
            "missing_field",
        )
        self.assertEqual(workbook["Run Metadata"]["A1"].value, "metadata_key")


if __name__ == "__main__":
    unittest.main()
