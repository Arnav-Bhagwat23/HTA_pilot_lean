from __future__ import annotations

import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from .schema import (
    ECONOMIC_EVALUATION_FIELDS,
    FIELD_LABELS,
    GUIDELINE_RESULT_FIELDS,
    HTA_RESULT_FIELDS,
    NMA_ITC_RESULT_FIELDS,
    TRIAL_RESULT_FIELDS,
)
from .storage import results_dir, slugify


OLD_PROJECT_SHEET_SPECS = [
    ("HTA Results", "hta_results", HTA_RESULT_FIELDS),
    ("Trial Results", "trial_results", TRIAL_RESULT_FIELDS),
    ("NMA Results", "nma_itc_results", NMA_ITC_RESULT_FIELDS),
    ("Economic Evaluation", "economic_evaluation", ECONOMIC_EVALUATION_FIELDS),
    ("Guideline Results", "guideline_results", GUIDELINE_RESULT_FIELDS),
]


def utc_timestamp() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def load_extraction_record(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def default_excel_path(record: dict[str, Any]) -> Path:
    document_set = record["document_set"]
    timestamp = utc_timestamp().replace(":", "-")
    return (
        results_dir()
        / "excel"
        / slugify(document_set["country"])
        / f"{slugify(document_set['product_name'])}__{timestamp}.xlsx"
    )


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _fallback_source_name(document: dict[str, Any]) -> str:
    source_name = document.get("source_name")
    if source_name:
        return source_name
    source_id = document.get("source_id") or ""
    labels = {
        "nice_uk": "NICE",
        "smc_uk": "Scottish Medicines Consortium",
        "has_france": "HAS",
        "gba_germany": "G-BA",
        "aifa_italy": "AIFA",
        "aemps_spain": "AEMPS",
        "pbac_australia": "PBAC",
    }
    return labels.get(source_id, source_id)


def _fallback_format(document: dict[str, Any]) -> str:
    file_format = document.get("format")
    if file_format:
        return file_format
    local_path = str(document.get("local_file_path") or "")
    document_url = str(document.get("document_url") or "")
    if local_path.lower().endswith(".pdf") or ".pdf" in document_url.lower():
        return "pdf"
    if document_url:
        return "html"
    return ""


def _field_value(field: dict[str, Any], key: str) -> Any:
    value = field.get(key)
    if isinstance(value, list):
        return "\n".join(str(item) for item in value)
    return value


def _field_label(field_name: str) -> str:
    return FIELD_LABELS.get(field_name, field_name.replace("_", " ").title())


def _extracted_value(field: dict[str, Any] | None) -> Any:
    if not field:
        return None
    return field.get("value")


def _old_project_headers(fields: tuple[str, ...]) -> list[str]:
    return [_field_label(field_name) for field_name in fields]


def build_old_project_rows(
    record: dict[str, Any], section_key: str, fields: tuple[str, ...]
) -> list[list[Any]]:
    if section_key == "hta_results":
        section = record.get(section_key, {})
        return [[_extracted_value(section.get(field_name)) for field_name in fields]]

    rows = []
    for item in record.get(section_key, []) or []:
        item_fields = item.get("fields", {})
        rows.append([_extracted_value(item_fields.get(field_name)) for field_name in fields])
    return rows


def iter_extracted_fields(record: dict[str, Any]):
    for field_name, field in record.get("hta_results", {}).items():
        yield "hta_results", None, None, field_name, field

    for section_key in (
        "trial_results",
        "nma_itc_results",
        "economic_evaluation",
        "guideline_results",
    ):
        for item in record.get(section_key, []) or []:
            for field_name, field in item.get("fields", {}).items():
                yield (
                    section_key,
                    item.get("row_id"),
                    item.get("row_label"),
                    field_name,
                    field,
                )


def _document_lookup(record: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        document.get("document_id"): document
        for document in record.get("document_set", {}).get("documents_considered", [])
        if document.get("document_id")
    }


def _used_fields_by_document(record: dict[str, Any]) -> dict[str, list[str]]:
    used: dict[str, list[str]] = {}
    for section_key, row_id, _row_label, field_name, field in iter_extracted_fields(record):
        document_id = field.get("source_document_id")
        if field.get("value") and document_id:
            field_path = f"{section_key}.{field_name}"
            if row_id:
                field_path = f"{section_key}[{row_id}].{field_name}"
            used.setdefault(document_id, []).append(field_path)
    return used


def _pages_by_document(record: dict[str, Any]) -> dict[str, list[str]]:
    pages: dict[str, list[str]] = {}
    for _section_key, _row_id, _row_label, _field_name, field in iter_extracted_fields(record):
        document_id = field.get("source_document_id")
        source_page = field.get("source_page")
        if field.get("value") and document_id and source_page:
            page_text = str(source_page)
            if page_text not in pages.setdefault(document_id, []):
                pages[document_id].append(page_text)
    return pages


def _confidence_summary(record: dict[str, Any]) -> dict[str, str]:
    confidence_rank = {"unknown": 0, "low": 1, "medium": 2, "high": 3}
    confidence_by_document: dict[str, list[str]] = {}
    for _section_key, _row_id, _row_label, _field_name, field in iter_extracted_fields(record):
        document_id = field.get("source_document_id")
        confidence = field.get("confidence")
        if field.get("value") and document_id and confidence:
            confidence_by_document.setdefault(document_id, []).append(confidence)

    summaries: dict[str, str] = {}
    for document_id, confidences in confidence_by_document.items():
        ordered = sorted(
            set(confidences), key=lambda item: confidence_rank.get(item, 0), reverse=True
        )
        summaries[document_id] = " / ".join(ordered)
    return summaries


def build_field_provenance_rows(record: dict[str, Any]) -> list[list[Any]]:
    document_set = record["document_set"]
    rows = []
    for section_key, row_id, row_label, field_name, field in iter_extracted_fields(record):
        rows.append(
            [
                document_set.get("product_name"),
                document_set.get("country"),
                section_key,
                row_id,
                row_label,
                field_name,
                _field_label(field_name),
                field.get("value"),
                bool(field.get("value")),
                field.get("fill_method"),
                field.get("confidence"),
                field.get("source_document_id"),
                field.get("source_document_title"),
                field.get("source_document_date"),
                field.get("source_document_url"),
                field.get("source_page"),
                field.get("evidence_snippet"),
                _field_value(field, "warnings"),
                "",
                "",
            ]
        )
    return rows


def build_documents_rows(record: dict[str, Any]) -> list[list[Any]]:
    document_set = record["document_set"]
    processed_document_ids = {
        entry.get("document_id") for entry in record.get("traceability", {}).get("audit_log", [])
    }
    rows = []
    for document in document_set.get("documents_considered", []):
        was_processed = document.get("document_id") in processed_document_ids
        skip_reason = ""
        if not was_processed:
            if not document.get("local_file_path"):
                skip_reason = "missing_local_file_or_non_pdf"
            else:
                skip_reason = "not_processed"
        rows.append(
            [
                document_set.get("product_name"),
                document_set.get("country"),
                document.get("processing_order"),
                document.get("document_id"),
                document.get("source_id"),
                _fallback_source_name(document),
                document.get("title"),
                document.get("document_type"),
                _fallback_format(document),
                document.get("event_date"),
                document.get("publication_date"),
                document.get("revision_date"),
                document.get("is_latest_version"),
                document.get("timeline_priority"),
                document.get("match_term"),
                document.get("match_confidence"),
                document.get("document_url"),
                document.get("local_file_path"),
                was_processed,
                skip_reason,
            ]
        )
    return rows


def build_audit_rows(record: dict[str, Any]) -> list[list[Any]]:
    document_set = record["document_set"]
    documents = _document_lookup(record)
    rows = []
    for index, entry in enumerate(record.get("traceability", {}).get("audit_log", []), start=1):
        attempted = entry.get("fields_attempted") or []
        filled = entry.get("fields_filled") or []
        not_filled = [field for field in attempted if field not in filled]
        document = documents.get(entry.get("document_id"), {})
        rows.append(
            [
                document_set.get("product_name"),
                document_set.get("country"),
                index,
                entry.get("timestamp"),
                entry.get("action"),
                entry.get("document_id"),
                document.get("title"),
                document.get("event_date"),
                _stringify(attempted),
                _stringify(filled),
                _stringify(not_filled),
                entry.get("notes"),
            ]
        )
    return rows


def build_warning_rows(record: dict[str, Any]) -> list[list[Any]]:
    document_set = record["document_set"]
    rows = []
    for section_key, row_id, row_label, field_name, field in iter_extracted_fields(record):
        warnings = field.get("warnings") or []
        confidence = field.get("confidence")
        issue_types: list[tuple[str, str]] = []
        if not field.get("value"):
            issue_types.append(("missing_field", "Field was not filled."))
        if confidence in {"low", "unknown"} and field.get("value"):
            issue_types.append((f"{confidence}_confidence", "Review confidence level."))
        issue_types.extend(("warning", str(warning)) for warning in warnings)

        for issue_type, warning_message in issue_types:
            rows.append(
                [
                    document_set.get("product_name"),
                    document_set.get("country"),
                    section_key,
                    row_id,
                    row_label,
                    field_name,
                    _field_label(field_name),
                    issue_type,
                    field.get("value"),
                    confidence,
                    warning_message,
                    field.get("source_document_id"),
                    field.get("source_document_title"),
                    field.get("source_page"),
                    "manual_review" if issue_type != "missing_field" else "accept_blank_or_rerun",
                    "",
                    "",
                ]
            )
    return rows


def build_metadata_rows(record: dict[str, Any], json_source_path: Path | None) -> list[list[Any]]:
    document_set = record["document_set"]
    traceability = record.get("traceability", {})
    progressive_fill = record.get("progressive_fill", {})
    extracted_fields = list(iter_extracted_fields(record))
    fields_total = len(extracted_fields)
    fields_filled = sum(1 for *_prefix, field in extracted_fields if field.get("value"))
    audit_document_ids = {
        entry.get("document_id") for entry in traceability.get("audit_log", [])
    }
    warnings_count = len(traceability.get("warnings") or [])
    warnings_count += sum(
        len(field.get("warnings") or []) for *_prefix, field in extracted_fields
    )
    values = {
        "schema_version": record.get("schema_version"),
        "product_name": document_set.get("product_name"),
        "country": document_set.get("country"),
        "indication": document_set.get("indication"),
        "created_at": traceability.get("created_at"),
        "updated_at": traceability.get("updated_at"),
        "extraction_model": traceability.get("extraction_model"),
        "extraction_status": traceability.get("extraction_status"),
        "allow_final_inference": progressive_fill.get("allow_final_inference"),
        "overwrite_populated_fields": progressive_fill.get("overwrite_populated_fields"),
        "missing_field_policy": progressive_fill.get("missing_field_policy"),
        "strategy": progressive_fill.get("strategy"),
        "documents_considered_count": len(document_set.get("documents_considered", [])),
        "documents_processed_count": len(audit_document_ids),
        "fields_total_count": fields_total,
        "fields_filled_count": fields_filled,
        "fields_missing_count": fields_total - fields_filled,
        "trial_rows_count": len(record.get("trial_results", []) or []),
        "nma_itc_rows_count": len(record.get("nma_itc_results", []) or []),
        "economic_evaluation_rows_count": len(record.get("economic_evaluation", []) or []),
        "guideline_rows_count": len(record.get("guideline_results", []) or []),
        "warnings_count": warnings_count,
        "json_source_path": str(json_source_path) if json_source_path else "",
        "excel_exported_at": utc_timestamp(),
    }
    return [[key, value] for key, value in values.items()]


def build_source_url_rows(record: dict[str, Any]) -> list[list[Any]]:
    document_set = record["document_set"]
    used_fields = _used_fields_by_document(record)
    pages = _pages_by_document(record)
    confidence = _confidence_summary(record)
    rows = []
    for document in document_set.get("documents_considered", []):
        document_id = document.get("document_id")
        rows.append(
            [
                document_set.get("product_name"),
                document_set.get("country"),
                document.get("source_id"),
                _fallback_source_name(document),
                document_id,
                document.get("title"),
                document.get("document_url"),
                document.get("local_file_path"),
                _stringify(used_fields.get(document_id, [])),
                _stringify(pages.get(document_id, [])),
                confidence.get(document_id, ""),
            ]
        )
    return rows


def write_extraction_excel(
    record: dict[str, Any], destination: Path, *, json_source_path: Path | None = None
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook = Workbook()
    workbook.remove(workbook.active)

    sheet_specs = [
        (
            sheet_name,
            _old_project_headers(fields),
            build_old_project_rows(record, section_key, fields),
        )
        for sheet_name, section_key, fields in OLD_PROJECT_SHEET_SPECS
    ]
    sheet_specs.extend(
        [
        (
            "Field Provenance",
            [
                "product_name",
                "country",
                "schema_section",
                "row_id",
                "row_label",
                "field_name",
                "field_label",
                "extracted_value",
                "is_filled",
                "fill_method",
                "confidence",
                "source_document_id",
                "source_document_title",
                "source_document_date",
                "source_document_url",
                "source_page",
                "evidence_snippet",
                "warnings",
                "review_status",
                "reviewer_notes",
            ],
            build_field_provenance_rows(record),
        ),
        (
            "Documents Considered",
            [
                "product_name",
                "country",
                "processing_order",
                "document_id",
                "source_id",
                "source_name",
                "title",
                "document_type",
                "format",
                "event_date",
                "publication_date",
                "revision_date",
                "is_latest_version",
                "timeline_priority",
                "match_term",
                "match_confidence",
                "document_url",
                "local_file_path",
                "was_processed_by_ai",
                "skip_reason",
            ],
            build_documents_rows(record),
        ),
        (
            "Extraction Audit Log",
            [
                "product_name",
                "country",
                "audit_step",
                "timestamp",
                "action",
                "document_id",
                "document_title",
                "document_date",
                "fields_attempted",
                "fields_filled",
                "fields_not_filled",
                "notes",
            ],
            build_audit_rows(record),
        ),
        (
            "Missing Fields & Warnings",
            [
                "product_name",
                "country",
                "schema_section",
                "row_id",
                "row_label",
                "field_name",
                "field_label",
                "issue_type",
                "current_value",
                "confidence",
                "warning_message",
                "source_document_id",
                "source_document_title",
                "source_page",
                "recommended_action",
                "reviewer_resolution",
                "reviewer_notes",
            ],
            build_warning_rows(record),
        ),
        ("Run Metadata", ["metadata_key", "metadata_value"], build_metadata_rows(record, json_source_path)),
        (
            "Source URLs",
            [
                "product_name",
                "country",
                "source_id",
                "source_name",
                "document_id",
                "title",
                "document_url",
                "local_file_path",
                "used_for_fields",
                "source_pages_used",
                "confidence_summary",
            ],
            build_source_url_rows(record),
        ),
        ]
    )

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    wrapped = Alignment(wrap_text=True, vertical="top")

    for sheet_name, headers, rows in sheet_specs:
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        sheet.append(headers)
        for row in rows:
            sheet.append(row)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = wrapped

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        if rows:
            table_ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
            table_name = "".join(ch for ch in sheet_name.title() if ch.isalnum())[:20]
            table = Table(displayName=f"{table_name}Table", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        else:
            sheet.append(["No rows extracted for this section."])
            sheet["A2"].fill = title_fill
            sheet["A2"].font = Font(italic=True)

        for column_index, header in enumerate(headers, start=1):
            values = [header]
            values.extend(
                _stringify(sheet.cell(row=row_index, column=column_index).value)
                for row_index in range(2, min(sheet.max_row, 30) + 1)
            )
            width = min(max(max(len(value) for value in values) + 2, 12), 60)
            sheet.column_dimensions[get_column_letter(column_index)].width = width

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def export_extraction_json_to_excel(
    json_path: Path, destination: Path | None = None
) -> Path:
    record = load_extraction_record(json_path)
    output_path = destination or default_excel_path(record)
    return write_extraction_excel(record, output_path, json_source_path=json_path)
